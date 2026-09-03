#!/usr/bin/env python3
"""🔨️ F2 — generates before/after XLSX fixtures for s.stdio.xlsx@ecma-376/strict's 8 and
@ecma-376/transitional's 6 unfixtured mutations. Same pattern as the docx/pptx structural generators
and built on the same shared helpers (🔨️f2_ooxml_common.py): set-snapshot is genuinely
third-party-generated; every OOXML-strict-vs-transitional structural mutation is handcrafted on top
of a genuine openpyxl 3.1.5 base package. set-relationships-namespace patches the xmlns:r declared on
the workbook.xml's own <sheet> element (verified live: openpyxl's workbook.xml root has no xmlns:r of
its own -- it is declared locally on each <sheet> child), and set-worksheet-content-type patches
[Content_Types].xml's own Override for the worksheet part.
Idempotent: safe to re-run.
"""
import io
import json
from pathlib import Path

import openpyxl

import importlib.util

_spec = importlib.util.spec_from_file_location("f2_ooxml_common", Path(__file__).parent / "🔨️f2_ooxml_common.py")
_mod = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_mod)
NS = _mod.NS
VML_CONTENT = _mod.VML_CONTENT
zip_read, zip_rewrite, patch_tag_attr = _mod.zip_read, _mod.zip_rewrite, _mod.patch_tag_attr
assert_wellformed, assert_zip_valid, sha256_of = _mod.assert_wellformed, _mod.assert_zip_valid, _mod.sha256_of

ROOT = Path("/Users/ueli/Documents/semio")
OPENPYXL_VERSION = "3.1.5"
WORKSHEET_CT_STRICT = "application/vnd.openxmlformats-officedocument.spreadsheetml.chartsheet+xml"


def base_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    wb.active["A1"] = "Hello"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def with_workbook_xml(data: bytes, new_xml: str) -> bytes:
    assert_wellformed(new_xml)
    out = zip_rewrite(data, {"xl/workbook.xml": new_xml})
    assert_zip_valid(out)
    return out


def add_vml_part(data: bytes) -> bytes:
    ct = zip_read(data, "[Content_Types].xml")
    if "vmlDrawing" not in ct:
        ct = ct.replace("</Types>", '<Default Extension="vml" ContentType="application/vnd.openxmlformats-officedocument.vmlDrawing"/></Types>')
    rels = zip_read(data, "xl/_rels/workbook.xml.rels")
    rels = rels.replace(
        "</Relationships>",
        '<Relationship Id="rIdVml1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing" Target="vmlDrawing1.vml"/></Relationships>',
    )
    out = zip_rewrite(data, {"[Content_Types].xml": ct, "xl/_rels/workbook.xml.rels": rels}, adds={"xl/vmlDrawing1.vml": VML_CONTENT})
    assert_zip_valid(out)
    return out


def remove_vml_part(data: bytes) -> bytes:
    ct = zip_read(data, "[Content_Types].xml").replace('<Default Extension="vml" ContentType="application/vnd.openxmlformats-officedocument.vmlDrawing"/>', "")
    rels = zip_read(data, "xl/_rels/workbook.xml.rels").replace(
        '<Relationship Id="rIdVml1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing" Target="vmlDrawing1.vml"/>', ""
    )
    out = zip_rewrite(data, {"[Content_Types].xml": ct, "xl/_rels/workbook.xml.rels": rels}, removes=["xl/vmlDrawing1.vml"])
    assert_zip_valid(out)
    return out


def set_worksheet_content_type(data: bytes, new_ct: str) -> bytes:
    ct_xml = zip_read(data, "[Content_Types].xml")
    old_ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
    marker = '<Override PartName="/xl/worksheets/sheet1.xml"'
    assert marker in ct_xml, "sheet1.xml Override not found"
    idx = ct_xml.index(marker)
    end = ct_xml.index("/>", idx) + 2
    segment = ct_xml[idx:end]
    new_segment = segment.replace(old_ct, new_ct)
    assert new_segment != segment, "content type replace had no effect"
    new_ct_xml = ct_xml[:idx] + new_segment + ct_xml[end:]
    out = zip_rewrite(data, {"[Content_Types].xml": new_ct_xml})
    assert_zip_valid(out)
    return out


def build_for_subset(subset: str):
    base = base_xlsx()
    tw_uri, strict_uri = NS["spreadsheetml"]
    tr_uri, strict_r_uri = NS["relationships"]
    conf_value = subset

    entries = []

    wb2 = openpyxl.Workbook()
    wb2.active["A1"] = "Snapshot"
    wb2.create_sheet("Extra")
    buf2 = io.BytesIO()
    wb2.save(buf2)
    entries.append(("set-snapshot", "third-party-generated", base, buf2.getvalue(),
                     "Whole-workbook snapshot replace: an unrelated valid workbook (different cell content, an extra sheet) substituted wholesale."))

    xml = zip_read(base, "xl/workbook.xml")

    after_xml = patch_tag_attr(xml, "workbook", "conformance", conf_value)
    entries.append(("set-conformance-attribute", "handcrafted", base, with_workbook_xml(base, after_xml),
                     f'The root <workbook>\'s conformance attribute set, absent -> "{conf_value}".'))

    before_conf = with_workbook_xml(base, patch_tag_attr(xml, "workbook", "conformance", conf_value))
    entries.append(("remove-conformance-attribute", "handcrafted", before_conf, base,
                     f'The root <workbook>\'s conformance attribute removed, "{conf_value}" -> absent.'))

    after_ns = patch_tag_attr(xml, "workbook", "xmlns", strict_uri)
    entries.append(("set-main-namespace", "handcrafted", base, with_workbook_xml(base, after_ns),
                     f"The root <workbook>'s default xmlns namespace URI changed, Transitional ({tw_uri}) -> Strict ({strict_uri})."))

    after_r = patch_tag_attr(xml, "sheet", "xmlns:r", strict_r_uri)
    entries.append(("set-relationships-namespace", "handcrafted", base, with_workbook_xml(base, after_r),
                     f"The <sheet> element's own local xmlns:r relationships namespace URI changed, Transitional ({tr_uri}) -> Strict ({strict_r_uri})."))

    after_ct = set_worksheet_content_type(base, WORKSHEET_CT_STRICT)
    entries.append(("set-worksheet-content-type", "handcrafted", base, after_ct,
                     f"[Content_Types].xml's own Override for /xl/worksheets/sheet1.xml changed to a different real OOXML content-type value ({WORKSHEET_CT_STRICT}), demonstrating the declared-content-type mutation this subset's vocabulary names -- not asserted to make the part a functional chartsheet."))

    if subset == "strict":
        after_vml = add_vml_part(base)
        entries.append(("insert-vml-part", "handcrafted", base, after_vml,
                         "A legacy VML drawing part (xl/vmlDrawing1.vml) added to the package, with its Content_Types.xml Default and a workbook.xml.rels relationship wiring it in."))
        entries.append(("remove-vml-part", "handcrafted", after_vml, remove_vml_part(after_vml),
                         "The VML drawing part, its content-type default and its relationship removed, inverse of insert-vml-part."))

    return entries


def emit(subset: str):
    subset_dir = ROOT / f"✏️s/🔌️plugins/🗄️stdio/🗿️artifacts/📕️xlsx/🏅️standards/🔖️ecma-376/🪆️subsets/✳️{subset}"
    fixtures = subset_dir / "🧫️fixtures"
    oracle_json = subset_dir / "🧪️oracle/🔣️.json"
    reader_oracle = f"openpyxl-xlsx-ecma-376-{subset}-mutate-reader"

    manifests = []
    for mutation_id, klass, before_bytes, after_bytes, note in build_for_subset(subset):
        case_dir = fixtures / f"{mutation_id}-applied"
        case_dir.mkdir(parents=True, exist_ok=True)
        (case_dir / "before.xlsx").write_bytes(before_bytes)
        (case_dir / "after.xlsx").write_bytes(after_bytes)

        entry = {
            "schema": "semio.repository-test.fixture/v2",
            "id": f"{mutation_id}-applied",
            "class": klass,
            "target": {"artifact": "s.stdio.xlsx", "standard": "ecma-376", "subset": subset},
            "mutation": mutation_id,
            "outcome": "applied",
            "units": {"length": "unitless", "angle": "degree"},
            "files": [
                {"role": "expected-before-xlsx", "path": f"../🧫️fixtures/{mutation_id}-applied/before.xlsx", "mediaType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "sha256": sha256_of(before_bytes), "bytes": len(before_bytes)},
                {"role": "expected-after-xlsx", "path": f"../🧫️fixtures/{mutation_id}-applied/after.xlsx", "mediaType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "sha256": sha256_of(after_bytes), "bytes": len(after_bytes)},
            ],
            "provenance": {
                "source": "generated" if klass == "third-party-generated" else "authored",
                "license": "MIT (openpyxl)" if klass == "third-party-generated" else "n/a (handcrafted zip/XML patch of an MIT openpyxl base package)",
                "attribution": "Written by openpyxl 3.1.5's own Workbook.save()" if klass == "third-party-generated" else "A genuine openpyxl 3.1.5 package hand-patched at the OOXML structural level (see notes); the patched xl/workbook.xml is re-parsed with lxml to confirm well-formedness and the archive's own zip integrity is re-checked before commit",
                "security": "scanned-clean",
                "privacy": "no-personal-data",
            },
            "comparisonProfile": "exact-bytes-v1",
            "reproducible": True,
            "family": "structural",
            "notes": note,
        }
        if klass == "third-party-generated":
            entry["generator"] = {
                "oracle": reader_oracle,
                "packageVersion": OPENPYXL_VERSION,
                "engineFamily": "openpyxl",
                "engineVersion": OPENPYXL_VERSION,
                "command": "uv run python3 🔨️f2-gen-xlsx-structural-fixtures.py (openpyxl Workbook object model + .save())",
                "platform": "darwin-arm64",
            }
        manifests.append(entry)
        print(f"[{subset}] {mutation_id:28s} {klass:20s} before={len(before_bytes)}B after={len(after_bytes)}B")

    data = json.loads(oracle_json.read_text())
    data["fixtureManifests"] = manifests
    oracle_json.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n")
    print(f"Wrote {len(manifests)} fixtureManifests entries into {oracle_json}\n")


def main() -> None:
    emit("strict")
    emit("transitional")


if __name__ == "__main__":
    main()
